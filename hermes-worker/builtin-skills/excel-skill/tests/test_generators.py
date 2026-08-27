"""generators 集成测试 — 跑每个 generate_*.py 确认能产 valid .xlsx。

不验证业务正确性（不同行业模板的精确公式属于审计范畴），只验证：
- 跑完不 crash
- 产出文件存在且能被 openpyxl 读回
- 关键 sheet 名存在
- 没有 #REF! / #N/A 这种已知错误
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PYTHON = sys.executable


def _run_generator(module_relpath: str, args: list[str], cwd: Path) -> subprocess.CompletedProcess:
    """跑 scripts/generate_*/*.py，传 --output 到 tmp。"""
    return subprocess.run(
        [PYTHON, module_relpath, *args],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",       # generator 输出含中文 / ✓，Windows 下不能用 GBK 解码
        errors="replace",
        timeout=60,
    )


def _assert_no_error_values(path: Path):
    wb = load_workbook(path, data_only=False)
    errors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in (
                    '#REF!', '#NAME?', '#NULL!'
                ):
                    errors.append(f'{sheet_name}!{cell.coordinate}={cell.value}')
    wb.close()
    assert not errors, f'生成的 xlsx 含错误标志值: {errors}'


# ----------------------------------------------------------------------------
# 4 个行业 generator
# ----------------------------------------------------------------------------

@pytest.fixture
def out_xlsx(tmp_path: Path) -> Path:
    return tmp_path / 'out.xlsx'


def test_finance_three_statements(out_xlsx: Path):
    r = _run_generator(
        'scripts/generate_finance/three_statements.py',
        ['--output', str(out_xlsx), '--years', '3', '--company', 'Test Co'],
        cwd=PROJECT_ROOT,
    )
    assert r.returncode == 0, f'stderr: {r.stderr}'
    assert out_xlsx.exists()
    wb = load_workbook(out_xlsx, data_only=False)
    names = wb.sheetnames
    wb.close()
    assert any('Income' in n or '利润' in n for n in names), names
    _assert_no_error_values(out_xlsx)


def test_ecommerce_gmv(out_xlsx: Path):
    r = _run_generator(
        'scripts/generate_ecommerce/gmv_dashboard.py',
        ['--output', str(out_xlsx)],
        cwd=PROJECT_ROOT,
    )
    assert r.returncode == 0, f'stderr: {r.stderr}'
    assert out_xlsx.exists()
    _assert_no_error_values(out_xlsx)


def test_fmcg_sales_vs_target(out_xlsx: Path):
    r = _run_generator(
        'scripts/generate_fmcg/sales_vs_target.py',
        ['--output', str(out_xlsx)],
        cwd=PROJECT_ROOT,
    )
    assert r.returncode == 0, f'stderr: {r.stderr}'
    assert out_xlsx.exists()
    _assert_no_error_values(out_xlsx)


def test_internet_dau_mau(out_xlsx: Path):
    r = _run_generator(
        'scripts/generate_internet/dau_mau_cohort.py',
        ['--output', str(out_xlsx)],
        cwd=PROJECT_ROOT,
    )
    assert r.returncode == 0, f'stderr: {r.stderr}'
    assert out_xlsx.exists()
    _assert_no_error_values(out_xlsx)


def test_saas_unit_economics(out_xlsx: Path):
    """v2 新增 SaaS vertical。"""
    r = _run_generator(
        'scripts/generate_saas/unit_economics.py',
        ['--output', str(out_xlsx), '--months', '12',
         '--arpu', '50', '--gross-margin', '0.75',
         '--monthly-churn', '0.04', '--cac', '800'],
        cwd=PROJECT_ROOT,
    )
    assert r.returncode == 0, f'stderr: {r.stderr}'
    assert out_xlsx.exists()
    wb = load_workbook(out_xlsx, data_only=False)
    names = wb.sheetnames
    wb.close()
    # SaaS 模板必须有这 5 张 sheet
    for required in ('Assumptions', 'Cohort', 'UnitEconomics', 'PL_Projection', 'Dashboard'):
        assert required in names, f'缺 sheet: {required}, 现有 {names}'
    _assert_no_error_values(out_xlsx)

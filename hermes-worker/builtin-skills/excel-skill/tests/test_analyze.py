"""analyze.py 测试 — 用合成 xlsx 验证检查项。"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.analyze import analyze_workbook


def _make_xlsx(tmp_path: Path, builder) -> Path:
    wb = Workbook()
    builder(wb)
    p = tmp_path / 'sample.xlsx'
    wb.save(p)
    wb.close()
    return p


# ----------------------------------------------------------------------------
# 各 check 项
# ----------------------------------------------------------------------------

def test_clean_workbook_passes(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws.title = 'Data'
        ws['A1'] = 100
        ws['A2'] = 200
        ws['A3'] = '=SUM(A1:A2)'

    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    assert r.passed, [f.code for f in r.errors]


def test_xa001_chinese_comma_in_formula(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws['A1'] = '=SUM(B1，B2)'  # 中文逗号
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    assert 'XA001' in {f.code for f in r.errors}


def test_xa002_error_value_in_cell(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws['A1'] = '#REF!'
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    assert 'XA002' in {f.code for f in r.errors}


def test_xa004_sheet_name_too_long(tmp_path: Path):
    def builder(wb):
        wb.active.title = 'a' * 32  # 32 字符
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    assert 'XA004' in {f.code for f in r.errors}


def test_xa007_magic_number_in_formula(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws['A1'] = '=B1*1234567'   # 大魔数
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    info_codes = {f.code for f in r.findings if f.severity == 'info'}
    assert 'XA007' in info_codes


def test_xa009_whole_column_reference(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws['A1'] = '=SUM(B:B)'
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    codes = {f.code for f in r.warnings}
    assert 'XA009' in codes


def test_xa008_merged_cells_detected(tmp_path: Path):
    """合并单元格应触发 XA008 info（回归：openpyxl 3.1+ 的 CellRange.size 是 dict，
    旧代码 `size > 1` 直接 TypeError 让整个 analyze 崩）。"""
    def builder(wb):
        ws = wb.active
        ws['A1'] = '标题'
        ws.merge_cells('A1:C1')
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)   # 不应抛 TypeError
    info_codes = {f.code for f in r.findings if f.severity == 'info'}
    assert 'XA008' in info_codes


def test_analyze_workbook_with_merges_does_not_crash(tmp_path: Path):
    """单格合并 / 多格合并混合，analyze_workbook 应正常跑完。"""
    def builder(wb):
        ws = wb.active
        ws.title = 'Data'
        ws['A1'] = 1
        ws.merge_cells('A1:A1')   # 退化单格
        ws['B1'] = 2
        ws.merge_cells('B1:D2')   # 真多格
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    assert r.to_dict()['summary']['total'] >= 1


def test_report_to_dict_has_summary(tmp_path: Path):
    def builder(wb):
        ws = wb.active
        ws['A1'] = '=SUM(A:A)'  # 触发 XA009 warning
    p = _make_xlsx(tmp_path, builder)
    r = analyze_workbook(p)
    d = r.to_dict()
    assert 'summary' in d
    assert d['summary']['warnings'] >= 1
    assert d['summary']['total'] >= 1


def test_nonexistent_file_raises(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        analyze_workbook(tmp_path / 'nope.xlsx')

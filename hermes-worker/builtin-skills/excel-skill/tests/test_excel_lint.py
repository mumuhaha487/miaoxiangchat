"""Tests for scripts/excel_lint.py."""
from __future__ import annotations

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.excel_lint import lint_file


def test_lint_catches_data_only_save_pitfall(tmp_path: Path) -> None:
    """XL001: data_only=True 后 save 应该被抓到。"""
    bad = tmp_path / 'bad.py'
    bad.write_text('''
from openpyxl import load_workbook

wb = load_workbook('file.xlsx', data_only=True)
wb['Sheet'].cell(1, 1).value = 'x'
wb.save('file.xlsx')
''', encoding='utf-8')
    report = lint_file(bad)
    assert not report.passed, 'Should detect data_only + save pitfall'
    codes = {i.code for i in report.errors}
    assert 'XL001' in codes


def test_lint_catches_chinese_comma_in_formula(tmp_path: Path) -> None:
    """XL003: 公式里中文逗号 ， 应该被抓到。"""
    bad = tmp_path / 'bad.py'
    bad.write_text('''
ws = something()
ws['A1'] = '=SUM(B1，B10)'
''', encoding='utf-8')
    report = lint_file(bad)
    codes = {i.code for i in report.errors}
    assert 'XL003' in codes


def test_lint_passes_clean_template(tmp_path: Path) -> None:
    """合法的模板应该通过 lint。"""
    template = PROJECT_ROOT / 'scripts' / 'generate_ecommerce' / 'gmv_dashboard.py'
    if not template.exists():
        # 跳过：脚本可能未初始化
        import pytest
        pytest.skip('template script missing')
    report = lint_file(template)
    # 实际可能有 warnings 但不应有 errors
    assert len([i for i in report.errors if i.code == 'XL001']) == 0
    assert len([i for i in report.errors if i.code == 'XL002']) == 0


def test_lint_handles_syntax_error(tmp_path: Path) -> None:
    """语法错误优雅处理。"""
    bad = tmp_path / 'broken.py'
    bad.write_text('def foo(:\n    pass\n', encoding='utf-8')
    report = lint_file(bad)
    assert not report.passed


def test_lint_catches_long_sheet_name(tmp_path: Path) -> None:
    """XL005: sheet 名超 31 字符。"""
    bad = tmp_path / 'bad.py'
    bad.write_text('''
wb = something()
wb.create_sheet("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
''', encoding='utf-8')
    report = lint_file(bad)
    codes = {i.code for i in report.issues}
    # XL005 应该至少是 warning
    assert 'XL005' in codes

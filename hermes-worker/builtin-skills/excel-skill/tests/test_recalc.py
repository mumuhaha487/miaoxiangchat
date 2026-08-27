"""recalc.py 测试 — 覆盖不依赖 LibreOffice 安装的可测路径。

完整重算路径需要本机装 LibreOffice，CI 不保证有；这里测：
- 找不到 LibreOffice 时优雅报错（exit 1），不抛异常
- 输入文件不存在时 exit 1
- find_libreoffice 的返回类型契约
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts import recalc


def _make_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 1
    ws['A2'] = 2
    ws['A3'] = '=SUM(A1:A2)'
    p = tmp_path / 'f.xlsx'
    wb.save(p)
    wb.close()
    return p


def test_missing_input_returns_1(tmp_path: Path):
    rc = recalc.main([str(tmp_path / 'nope.xlsx')])
    assert rc == 1


def test_no_libreoffice_degrades_gracefully(tmp_path: Path, monkeypatch):
    """模拟未装 LibreOffice：应 exit 1 且不抛异常。"""
    monkeypatch.setattr(recalc, 'find_libreoffice', lambda: None)
    p = _make_xlsx(tmp_path)
    rc = recalc.main([str(p)])
    assert rc == 1
    # 原文件未被破坏
    assert p.exists()


def test_find_libreoffice_returns_str_or_none():
    result = recalc.find_libreoffice()
    assert result is None or isinstance(result, str)
